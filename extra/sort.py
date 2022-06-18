from openpyxl import load_workbook
import random, copy

def sort(book_list,n):
    i=0
    my_list=book_list[:]
    print("************************\n")
    print("Here is the random list\n")
    while(i!=n):
        random_ch=random.choice(my_list)
        print(random_ch)
        my_list.remove(random_ch)        
        i+=1
    print("************************\n")
    

print("You want to:\n1. Import excel\n2. Enter manually")
x = int(input())
#print(x+"\t"+type(x))
if x==1:
    print("Importing")
    wb = load_workbook("goodreads_library_export.xlsx")  # Work Book
    ws = wb.get_sheet_by_name('TBR')  # Work Sheet
    column = ws['A']  # Column
    column_list = [column[x].value for x in range(len(column))]
    sort(column_list, int(input("How many random books do you want?: ")))
    #print(column_list)

elif x==2:
    print("Enter the choices (enter x to exit):")
    books=[]
    ch='y'
    while(ch!='x'):
        ch=input("Book Title: ")
        if ch!='x':
            books.append(ch)
    print("Your list is:\n",books)
    sort(books,len(books),"\n")
    while(input("Do you want to sort again? (Y/N)").lower() != 'n'):
        sort(books,len(books))
    
        
#import pandas as pd

#print("Importing")
#df = pd.read_excel('goodreads_library_export.xlsx') # can also index sheet by name or fetch all sheets
#mylist = df['A'].tolist()

#print(mylist)
